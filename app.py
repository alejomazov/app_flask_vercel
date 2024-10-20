from flask import Flask, request, send_file, render_template
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Alignment
import os

def deploy_render():
    app = Flask(__name__)

    def procesar_tipo_1(file_path, output_file):
        # Leer el CSV y hacer modificaciones específicas del tipo 1
        file_to_modificate = pd.read_csv(file_path, decimal=".")
        if file_to_modificate.columns.values[3]== " PENDIENTE":
            print("este es ws12")
            modificate_file = file_to_modificate.reindex(
                [' PENDIENTE', ' DIRECCIóN DE PENDIENTE', 'X', ' Y', ' Z', ' RUMBO', ' LONGITUD', ' ÁREA'], axis=1)
            modificate_file = modificate_file.rename(columns={
                ' PENDIENTE': 'BUZAMIENTO',
                ' DIRECCIóN DE PENDIENTE': "DIRECCIÓN DE INCLINACIÓN",
                ' Z': 'Z',
                ' RUMBO': 'RUMBO',
                ' LONGITUD': 'LONGITUD (m)',
                ' ÁREA': 'ÁREA (m)'})
            modificate_file["PERSISTENCIA (m)"] = modificate_file["LONGITUD (m)"].apply(
                lambda x: "<1" if x < 1 else "1 a 3" if x < 3 else "3 a 10" if x < 10 else "10 a 20" if x < 20 else ">20")

        else:
            print("este es ws07 o ws08")
            modificate_file = file_to_modificate.reindex(
                [' BUZAMIENTO', ' DIRECCIóN DE INCLINACIóN', 'X', ' Y', ' Z', ' RUMBO', ' LONGITUD', ' ÁREA'], axis=1)
            modificate_file = modificate_file.rename(columns={
                ' BUZAMIENTO': 'BUZAMIENTO',
                ' DIRECCIóN DE INCLINACIóN': "DIRECCIÓN DE INCLINACIÓN",
                ' Z': 'Z',
                ' RUMBO': 'RUMBO',
                ' LONGITUD': 'LONGITUD (m)',
                ' ÁREA': 'ÁREA (m)'})
            modificate_file["PERSISTENCIA (m)"] = modificate_file["LONGITUD (m)"].apply(
                lambda x: "<1" if x < 1 else "1 a 3" if x < 3 else "3 a 10" if x < 10 else "10 a 20" if x < 20 else ">20")

        modificate_file.to_excel(output_file, index=False, float_format="%.3f")

        # Ajustar estilos en el Excel
        workbook = load_workbook(output_file)
        worksheet = workbook.active
        for cell in worksheet[1]:
            cell.font = cell.font.copy(bold=False)
            cell.border = Border()
            cell.alignment = Alignment(horizontal='left')

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = max_length + 2

        workbook.save(output_file)

    def procesar_tipo_2(file_path, output_file):
        # Lógica para procesar archivos del Tipo 2
        # Modificar este código según las necesidades del Tipo 2
        file_to_modificate = pd.read_csv(file_path, decimal=".")
        modificate_file = file_to_modificate.reindex(
                [' BUZAMIENTO', ' DIRECCIóN DE INCLINACIóN', 'X', ' Y', ' Z', ' RUMBO', ' LONGITUD', ' ÁREA', ' NOMBRE DE OBJETO'], axis=1)
        modificate_file = modificate_file.rename(columns={
            ' BUZAMIENTO': 'BUZAMIENTO',
            ' DIRECCIóN DE INCLINACIóN': "DIRECCIÓN DE INCLINACIÓN",
            ' Z': 'Z',
            ' RUMBO': 'RUMBO',
            ' LONGITUD': 'LONGITUD (m)',
            ' ÁREA': 'ÁREA (m)',
            ' NOMBRE DE OBJETO':'NOMBRE DE OBJETO'})
        #Agregamos la columna GRUPO, buscamos a la columna que pertence, extraemos el numero y lo agregamos 
        # a grupo como un entero
        modificate_file['GRUPO'] = modificate_file['NOMBRE DE OBJETO'].str.extract(r'(\d+)').astype(int)
        
        # Extraer el número y la letra después del guion bajo usando expresiones regulares
        modificate_file['NUMERO'] = modificate_file['NOMBRE DE OBJETO'].str.extract(r'EST_(\d+)').astype(int)
        modificate_file['LETRA'] = modificate_file['NOMBRE DE OBJETO'].str.extract(r'EST_\d+([a-zA-Z])')

        # Ordenar por 'NUMERO' y luego por 'LETRA'
        modificate_file = modificate_file.sort_values(by=['NUMERO', 'LETRA'])

        # Eliminar las columnas auxiliares si ya no son necesarias
        modificate_file = modificate_file.drop(columns=['NUMERO', 'LETRA'])

        # Reiniciar el índice para que sea secuencial
        modificate_file = modificate_file.reset_index(drop=True)
        
        # Redondear las columnas 'BUZAMIENTO' y 'DIRECCIÓN DE INCLINACIÓN' a 0 decimales
        modificate_file[['BUZAMIENTO', 'DIRECCIÓN DE INCLINACIÓN']] = modificate_file[['BUZAMIENTO', 'DIRECCIÓN DE INCLINACIÓN']].round(0)

        modificate_file.to_excel(output_file, index=False, float_format="%.2f")

        # Ajustar estilos en el Excel
        workbook = load_workbook(output_file)
        worksheet = workbook.active
        for cell in worksheet[1]:
            cell.font = cell.font.copy(bold=False)
            cell.border = Border()
            cell.alignment = Alignment(horizontal='left')

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = max_length + 2

        workbook.save(output_file)

    def procesar_tipo_3(file_path, output_file_tangram):
        # Lógica para procesar archivos del Tipo 3
        # Modificar este código según las necesidades del Tipo 3
        file_to_modificate = pd.read_csv(file_path, decimal=".")
        modificate_file = file_to_modificate.rename(columns={"X":"Este",
                                                             " Y": "Norte",
                                                             " Z":"Cota"," BUZAMIENTO":"Dip",
                                                             " DIRECCIóN DE INCLINACIóN":"Dip Direction",
                                                             " LONGITUD":"Radio"})
        num_row = len(modificate_file)
        modificate_file['Tipo']=("Disco")
        modificate_file["ID"] = range(1, num_row + 1) #asigna el id sucesivamente
        modificate_file = modificate_file.reindex(["ID","Este","Norte","Cota","Tipo","Dip","Dip Direction","Radio"], axis=1)
        
        modificate_file.to_csv(output_file_tangram, index=False,sep=",")
        

    def procesar_tipo_4(file_path, output_file_tangram):
        # Lógica para procesar archivos del Tipo 4
        # Modificar este código según las necesidades del Tipo 4
        file_to_modificate = pd.read_csv(file_path, decimal=".")
        modificate_file = file_to_modificate.rename(columns={"X":"Este",
                                                             " Y": "Norte",
                                                             " Z":"Cota"," BUZAMIENTO":"Dip",
                                                             " DIRECCIóN DE INCLINACIóN":"Dip Direction",
                                                             " LONGITUD":"Radio"})
        num_row = len(modificate_file)
        modificate_file['Tipo']=("Otro")
        modificate_file["ID"] = range(1, num_row + 1) #asigna el id sucesivamente
        modificate_file = modificate_file.reindex(["ID","Este","Norte","Cota","Tipo","Dip","Dip Direction","Radio"], axis=1)
        
        modificate_file.to_csv(output_file_tangram, index=False,sep=",")
        pass

    @app.route('/')
    def home():
        return render_template('upload.html')

    @app.route('/upload', methods=['POST'])
    def upload_file():
        file = request.files['file']
        tipo = request.form.get('tipo')
        nombre_descarga = request.form.get('nombre_descarga', 'archivo_procesado')

        if file and tipo:
            # Asegurarse de que la carpeta 'uploads' existe
            os.makedirs('uploads', exist_ok=True)
            
            # Guardar el archivo subido temporalmente
            file_path = os.path.join('uploads', file.filename)
            file.save(file_path)
            
            # Definir el nombre de salida para el archivo procesado
            output_file = f"uploads/{nombre_descarga}.xlsx"
            output_file_tangram = f"uploads/{nombre_descarga}.csv"

            # Procesar el archivo según el tipo seleccionado
            if tipo == "1":
                procesar_tipo_1(file_path, output_file)
            elif tipo == "2":
                procesar_tipo_2(file_path, output_file)
            elif tipo == "3":
                procesar_tipo_3(file_path, output_file_tangram)
                return send_file(output_file_tangram, as_attachment=True, download_name=f"{nombre_descarga}.csv")
                
            elif tipo == "4":
                procesar_tipo_4(file_path, output_file_tangram)
                return send_file(output_file_tangram, as_attachment=True, download_name=f"{nombre_descarga}.csv")

            # Eliminar el archivo subido para mantener el directorio limpio
            os.remove(file_path)

            # Enviar el archivo procesado de vuelta al cliente para su descarga
            return send_file(output_file, as_attachment=True, download_name=f"{nombre_descarga}.xlsx")

        return "Error: Archivo o tipo no válido"
    return app
if __name__ == "__main__":
    app = deploy_render()
    app.run()
